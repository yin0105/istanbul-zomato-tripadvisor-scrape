import sys, time, xlsxwriter, os, openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent
from selenium.webdriver.common.action_chains import ActionChains
from os.path import join, dirname
from datetime import datetime
from openpyxl import load_workbook
from threading import Thread
from os.path import join, dirname, realpath
# from dotenv import load_dotenv


class ZomatoThread(Thread):
 
    def __init__(self, locality_name, cur_row, url):
        Thread.__init__(self)
        self.locality_name = locality_name
        self.cur_row = cur_row
        self.url = url
         
    def run(self):
        global file_open_flag
        global details
        start_time = datetime.now()
        ua = UserAgent()
        userAgent = ua.random
        userAgent = userAgent.split(" ")
        userAgent[0] = "Mozilla/5.0"
        userAgent = " ".join(userAgent)
        print("userAgent = " + userAgent)
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('user-agent={0}'.format(userAgent))
        chrome_options.add_argument("--headless")
        # chrome_options.add_argument("window-size=1280,800")

        chrome_options.add_argument('--log-level=0')
        path = join(dirname(__file__), 'webdriver', 'chromedriver.exe')
        driver = webdriver.Chrome (executable_path = path, options = chrome_options )
        # driver.maximize_window()
        driver.get(self.url)

        while True:
            try:
                rest_name = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//main/div/section[3]/section/section[1]/div/h1"))).text
                break
            except:
                time.sleep(0.1)
                pass
        
        company = driver.find_element_by_xpath("//main/div/section[3]/section/section[1]/section[1]/a").text
        rating = ''
        commeters = ''
        try:
            rating = driver.find_element_by_xpath("//main/div/section[3]/section/section[2]/section/div[1]/p").text
            commeters = driver.find_element_by_xpath("//main/div/section[3]/section/section[2]/section/div[2]/p").text
            commeters = driver.split(" ")[0]
        except:
            pass
        cuisine = ", ".join([e.text for e in driver.find_elements_by_xpath("//h3[contains(text(), 'Cuisines')]/following-sibling::section/a")])
        cost = ''
        cost_alcohol = ''
        try:
            cost_1 = driver.find_element_by_xpath("//h3[contains(text(), 'Average Cost')]/following-sibling::p[1]").text
            if cost_1.find("with alcohol") > -1:
                cost_alcohol = cost_1
            else:
                cost = cost_1
        except:
            pass

        try:
            cost_2 = driver.find_element_by_xpath("//h3[contains(text(), 'Average Cost')]/following-sibling::p[3]").text
            if cost_2.find("₺") > -1:
                if cost_2.find("with alcohol") > -1:
                    cost_alcohol = cost_2
                else:
                    cost = cost_2
        except:
            pass

        address = driver.find_element_by_xpath("//h5[contains(text(), 'Direction')]/following-sibling::p").text
        opening_hours = ''
        try:
            opening_hours = driver.find_element_by_xpath("//span[contains(text(), 'Close')]/following-sibling::span[1]").text
        except:
            try:
                opening_hours = driver.find_element_by_xpath("//span[contains(text(), 'Open')]/following-sibling::span[1]").text
            except:
                pass
        # xlsfile_name = "xls\\zomato\\" + self.locality_name + ".xlsx"
        # wb = load_workbook(xlsfile_name)
        # ws = wb.active
        # ws.cell(row=self.cur_row, column=2).value = rest_name
        # ws.cell(row=self.cur_row, column=3).value = rating
        # ws.cell(row=self.cur_row, column=4).value = commeters
        # ws.cell(row=self.cur_row, column=5).value = cuisine
        # ws.cell(row=self.cur_row, column=6).value = cost_alcohol
        # ws.cell(row=self.cur_row, column=7).value = cost
        # ws.cell(row=self.cur_row, column=8).value = address
        # ws.cell(row=self.cur_row, column=10).value = opening_hours
        details[self.cur_row] = {"rest_name": rest_name, "rating": rating, "commeters": commeters, "cuisine": cuisine, "cost_alcohol": cost_alcohol, "cost": cost, "address": address, "opening_hours": opening_hours, "company": company}
        # wb.save(xlsfile_name)
        # wb.close()
        # print("wrote at row " + str(self.cur_row), "  :: address=", address, ", opening_hours=", opening_hours)
        driver.quit()


def get_urls(locality_index):     
    cur_row = 1
    start_time = datetime.now()
    ua = UserAgent()
    userAgent = ua.random
    userAgent = userAgent.split(" ")
    # userAgent[0] = "Mozilla/5.0"
    userAgent = " ".join(userAgent)
    print("userAgent = " + userAgent)
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('user-agent={0}'.format(userAgent))
    chrome_options.add_argument("--headless")
    # chrome_options.add_argument("window-size=1280,800")

    chrome_options.add_argument('--log-level=0')
    path = join(dirname(__file__), 'webdriver', 'chromedriver.exe')
    driver = webdriver.Chrome (executable_path = path, options = chrome_options )
    # driver.maximize_window()

    #Remove navigator.webdriver Flag using JavaScript
    # driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.get("https://www.zomato.com/istanbul")

    while True:
        try:
            # popular_localities = driver.find_elements_by_xpath("//div[@class='title']/following-sibling::div/div/div")
            popular_localities = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='title']/following-sibling::div/div/div")))
            print("get Popular localities: ", popular_localities)
            break
        except:
            time.sleep(0.1)
            pass
    popular_localities_url = []
    popular_localities_name = []
    for locality in popular_localities:
        popular_localities_url.append(locality.find_element_by_xpath("./a").get_attribute("href"))
        locality_name = locality.find_element_by_xpath("./a/div/div/h5").text
        locality_name= locality_name[:locality_name.find(" (")]
        popular_localities_name.append(locality_name)
    print("get Popular localities URL")
    print("popular_localities_url = ", popular_localities_url)

    # for locality_name, locality_url in zip(popular_localities_name, popular_localities_url):
    
    locality_name = popular_localities_name[locality_index]
    locality_url = popular_localities_url[locality_index]
    xlsfile_name = "xls\\zomato\\" + locality_name + ".xlsx"
    workbook = xlsxwriter.Workbook(xlsfile_name)
    worksheet = workbook.add_worksheet()  
    col_headers = ['No', 'Restaurant', 'Rating', 'Number of Comments', 'Cuisine', 'Cost with alcohol', 'Cost without alcohol', 'Address', 'Address as coordinate', 'Opening hours', 'Locality', 'URL', 'Image URL']
    
    for c, i in zip(col_headers, range(len(col_headers))):
        worksheet.write(0, i, c)

    print("time = ",(datetime.now() - start_time).total_seconds())
    print("locality_name = ", locality_name)
    # rest_data = {}
    rest_urls = []
    img_urls = []
    count = 0
    
    driver.get(locality_url) # + "?sort=cd&category=2")
    print("go to new page '" + locality_url + "'")
    # while True:
    #     try:
    #         categories = driver.find_elements_by_xpath("(//section[@role='tablist'])[1]/div")
    #         print("get categories")
    #         break
    #     except:
    #         time.sleep(0.1)
    #         pass
    # is_dining_out = True
    # time.sleep(0.5)
    # print("len of categories = ", len(categories))
    # for category in categories:
    #     if not is_dining_out:
    #         pass
    restaurants = driver.find_elements_by_xpath("//div[@class='jumbo-tracker']")
    restaurants_count = len(restaurants)
    print(restaurants_count)
    restaurants_pre_count = 0

    while restaurants_pre_count != restaurants_count:        
        driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
        ok = False
        restaurants_pre_count = restaurants_count
        restaurants_pre_count_2 = restaurants_count
        for i in range(5):            
            time.sleep(1)
            restaurants = driver.find_elements_by_xpath("//div[@class='jumbo-tracker']")
            restaurants_count = len(restaurants)
            if restaurants_pre_count_2 != restaurants_count: 
                restaurants_pre_count_2 = restaurants_count
                ok = True
            elif ok:
                break
        print("pre_count = " + str(restaurants_pre_count) + "  count = " + str(restaurants_count))
    
    try:
        end_of_results = driver.find_element_by_xpath("//h3[contains(text(), 'End of search results')]")
        print("####  End of search results  ####")
    except:
        pass
    
    cur_time = datetime.now()
    
    for restaurant in restaurants:
        # //div[@class='jumbo-tracker']/div/a/div/img
        rest_url = restaurant.find_element_by_xpath("./div/a[1]").get_attribute("href")
        img_url = restaurant.find_element_by_xpath("./div/a[1]/div[1]/img").get_attribute("src")
        
        if not rest_url in rest_urls:
            print("rest_url = ", rest_url)
            rest_urls.append(rest_url)
            img_urls.append(img_url)
            worksheet.write(cur_row, 0, cur_row)
            worksheet.write(cur_row, 10, locality_name)
            worksheet.write(cur_row, 11, rest_url)
            # worksheet.write(cur_row, 12, img_url)
            cur_row += 1
    
    workbook.close()

def get_details(locality_name, start_row, time_interval):
    global details
    xlsfile_name = "xls\\zomato\\" + locality_name + ".xlsx"
    wb = load_workbook(xlsfile_name)
    ws = wb.active    
    total_rows = 2
    rest_urls = []
    while True:    
        if ws.cell(row=total_rows, column=1).value == None: break
        url = ws.cell(row=total_rows, column=12).value
        rest_urls.append(url)
        total_rows += 1
    total_rows -= 1
    

    for i in range(start_row + 1, total_rows + 1):
        t = ZomatoThread(locality_name, i, rest_urls[i - 2])
        t.start()
        time.sleep(time_interval)
        
        print(len(details))

    while True:
        print(len(details))
        if len(details) == total_rows - start_row:
            for row in details:
                ws.cell(row=row, column=2).value = details[row].rest_name
                ws.cell(row=row, column=3).value = details[row].rating
                ws.cell(row=row, column=4).value = details[row].commeters
                ws.cell(row=row, column=5).value = details[row].cuisine
                ws.cell(row=row, column=6).value = details[row].cost_alcohol
                ws.cell(row=row, column=7).value = details[row].cost
                ws.cell(row=row, column=8).value = details[row].address
                ws.cell(row=row, column=9).value = details[row].company
                ws.cell(row=row, column=10).value = details[row].opening_hours

        time.sleep(time_interval)


def excel_merge():
    xlsfile_name = "xls\\zomato\\total.xlsx"
    # wb = load_workbook(xlsfile_name)
    wb = openpyxl.Workbook()
    ws = wb.active
    cur_row = 1

    directory = join(dirname(realpath(__file__)), "xls", "zomato")
    print("=##", directory, "##")
    urls = []
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"): 
            wb_2 = load_workbook(join(directory, filename))
            ws_2 = wb_2.active
            if cur_row == 1:
                for c in range(1, 13):
                    ws.cell(row=1, column=c).value = ws_2.cell(row=1, column=c).value
                cur_row = 2
            cur_row_2 = 2
            while True:    
                if ws_2.cell(row=cur_row_2, column=1).value == None: break
                url = ws_2.cell(row=cur_row_2, column=12).value
                if not url in urls:
                    urls.append(url)

                    ws.cell(row=cur_row, column=1).value = cur_row - 1
                    for c in range(2, 13):
                        ws.cell(row=cur_row, column=c).value = ws_2.cell(row=cur_row_2, column=c).value
                    print(cur_row)
                    cur_row = cur_row + 1

                cur_row_2 += 1
    wb.save(xlsfile_name)
    wb.close()
        
    # total_rows = 2
    # rest_urls = []
    # while True:    
    #     if ws.cell(row=total_rows, column=1).value == None: break
    #     url = ws.cell(row=total_rows, column=12).value
    #     rest_urls.append(url)
    #     total_rows += 1
    # total_rows -= 1
    

    # for i in range(start_row + 1, total_rows + 1):
    #     t = ZomatoThread(locality_name, i, rest_urls[i - 2])
    #     t.start()
    #     time.sleep(time_interval)
        
    #     print(len(details))

    # while True:
    #     print(len(details))
    #     if len(details == tatal_rows - start_row):
    #         for row in details:
    #             ws.cell(row=row, column=2).value = details[row].rest_name
    #             ws.cell(row=row, column=3).value = details[row].rating
    #             ws.cell(row=row, column=4).value = details[row].commeters
    #             ws.cell(row=row, column=5).value = details[row].cuisine
    #             ws.cell(row=row, column=6).value = details[row].cost_alcohol
    #             ws.cell(row=row, column=7).value = details[row].cost
    #             ws.cell(row=row, column=8).value = details[row].address
    #             ws.cell(row=row, column=9).value = details[row].company
    #             ws.cell(row=row, column=10).value = details[row].opening_hours

    #     time.sleep(time_interval)


        # rest_name = restaurant.find_element_by_xpath("./div/a[2]/p[1]").text
        # print("Consuming Time: 2", (datetime.now() - start_time).total_seconds())
        # if rest_name in rest_data: continue
        # rest_rating = ''
        # rest_comments = ''
        # try:
        #     rest_rating = restaurant.find_element_by_xpath("./div/a[2]/div/section/div[1]/p").text
        #     rest_comments = restaurant.find_element_by_xpath("./div/a[2]/div/section/div[2]/p").text
        #     rest_comments = rest_comments.split(" ")[0][1:]
        # except:
        #     pass
        # print("Consuming Time: 3", (datetime.now() - start_time).total_seconds())
        
        # rest_cuisine = restaurant.find_element_by_xpath("./div/a[2]/p[2]").text
        # print("Consuming Time: 4", (datetime.now() - start_time).total_seconds())
        # rest_budget = restaurant.find_element_by_xpath("./div/a[2]/p[4]/span").text
        # print("Consuming Time: 5", (datetime.now() - start_time).total_seconds())
        # rest_data[rest_name] = {'rating': rest_rating, 'comments': rest_comments, 'cuisine': rest_cuisine, 'budget': rest_budget}
        # count += 1
        # print(count, rest_name, rest_rating, rest_comments, rest_cuisine, rest_budget)

        # is_dining_out = False

    # for rest_url, img_url in zip(rest_urls, img_urls):
           
    


details = {}
file_open_flag = False
if sys.argv[1] == "url":
    get_urls(int(sys.argv[2]))   
elif sys.argv[1] == "detail":
    get_details(sys.argv[2], int(sys.argv[3]), int(sys.argv[4]) )
elif sys.argv[1] == "merge":
    excel_merge()

# End of search results h3